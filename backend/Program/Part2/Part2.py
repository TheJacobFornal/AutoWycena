from pathlib import Path
import os
from openpyxl import load_workbook



curr_dir = Path(__file__).parent
#main_dir = curr_dir.parent.parent
#excel_dir = main_dir / "Excel"
#excel_estimate_path = excel_dir / "_kalkulacja edit.xlsx"


def find_new_row():
    for row in range (8, 100):
        empty = True
        for col in range(5, 8):
            if ws.cell(row, col).value != None:
                empty = False
                break
        if empty:
            return row
        else:
            continue
    return None


def insert_date(row, task_table):
    for elem in task_table:
        task = elem[0]
        value = elem[1]
        if isinstance(value, str):
            value = value.strip()
        ws.cell(row, task).value = value


def insert_id(elem_id, row):
    elem_id = elem_id.strip()
    ws.cell(row, 5).value = elem_id

def main(elem_id, task_table, excel_estimate_path):
    global ws
    wb = load_workbook(excel_estimate_path)
    ws = wb.active


    row = find_new_row()

    insert_date(row, task_table)

    insert_id(elem_id, row)

    wb.save(excel_estimate_path)


