import datetime
from openpyxl import load_workbook
from pathlib import Path
import re
from Program.Part2 import Part2 as part_2

curr_dir = Path(__file__).parent
main_dir = curr_dir.parent.parent
excel_dir = main_dir / "Excel"


orders_path = None
excel_estimate_path = None

def get_month_sheet():
    months = [
        "Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
        "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"
    ]

    month_sheet = [name for name in wb.sheetnames if any(name.startswith(m) for m in months)]

    return month_sheet

Excel_map_1 = {
    5: 6,
    6: 7,
    7: 21,
    8: 36,
    9: 25,
    10: 28,
    11: 32,
    12: 34,
    13: 30,
    14: 38,

    18: 8,
    19: 13,
    20: 14,
    21: 15,
    22: 11,
    23: 12
}

Excel_map_2 = {
    5: 6,
    6: 7,
    7: 21,
    8: 36,
    9: 25,
    10: 28,
    11: 32,
    12: 34,
    13: 30,
    14 : 100,                                   ## jaki numer ma spawanie
    15: 38,

    19: 8,
    20: 13,
    21: 14,
    22: 15,
    23: 11,
    24: 12
}

task_table = []

def newer_Excel(ws):
    name = ws.title
    match = re.search(r'\d+', name)

    if match:
        number = int(match.group())
        if number >= 25:
            return True
        else:
            return False
    return None

def add_to_table(ws, row, col):
    value = ws.cell(row, col).value

    if newer_Excel(ws):
        elem_col = Excel_map_2[col]
    else:
        elem_col = Excel_map_1[col]

    task_table.append([elem_col, value])


def look_for_operation(ws, row):                                #go through cols in row
    if newer_Excel(ws):
        for col in range(7, 16):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                add_to_table(ws, row, col)
    else:
        for col in range(7, 15):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                add_to_table(ws, row, col)


def check_if_canceled(ws, row):
    cell_value = ws.cell(row=row, column=1).value
    if isinstance(cell_value, datetime.datetime):
        return False
    else:
        return True


def get_name_quantity(ws, row):
    #task_table.append(["###", "###"])
    name_col = 5
    quantity_col = 6

    add_to_table(ws, row, name_col)
    add_to_table(ws, row, quantity_col)

    if newer_Excel(ws):
        gatunek_col = 19
    else:
        gatunek_col = 18

    if ws.cell(row, gatunek_col).value is not None:
        add_to_table(ws, row, gatunek_col)
        print(gatunek_col)

def get_dimenstions(ws, row):
    if newer_Excel(ws):
        a_col = 20
        b_col = 21
        c_col = 22
        diameter_col = 23
        length_col = 24
    else:
        a_col = 19
        b_col = 20
        c_col = 21
        diameter_col = 22
        length_col = 23

    a = ws.cell(row, a_col)
    b = ws.cell(row, b_col)
    c = ws.cell(row, c_col)

    diameter = ws.cell(row, diameter_col)
    length = ws.cell(row, length_col)

    if all(cell.value is not None for cell in (a, b, c)):
        add_to_table(ws, row, a_col)
        add_to_table(ws, row, b_col)
        add_to_table(ws, row, c_col)

    elif all(cell.value is not None for cell in (diameter,length)):
        add_to_table(ws, row, diameter_col)
        add_to_table(ws, row, length_col)


def get_line(target_id, number_of_elem): 
    global excel_estimate_path
    counter = 0
    found_any = False
    finish = False

    for sheet in reversed(get_month_sheet()):
        ws = wb[sheet]
        if finish:
            break

        for row in range(2, ws.max_row + 1):
            if int(counter) >= int(number_of_elem):
                finish = True
                break

            cell_value = ws.cell(row=row, column=4).value
            value = str(cell_value).strip()
            if str(value) == target_id:
                if check_if_canceled(ws, row):
                    continue
                newer_Excel(ws)
                get_name_quantity(ws, row)
                get_dimenstions(ws, row)
                look_for_operation(ws, row)
                part_2.main(target_id, task_table, excel_estimate_path)
                task_table.clear()
                counter += 1
                found_any = True

    return found_any

       





def main(index, orders_path_f, excel_estimate_path_f, number_of_elem):
    global orders_path, excel_estimate_path, wb
    orders_path = orders_path_f
    excel_estimate_path = excel_estimate_path_f

    wb = load_workbook(orders_path)
    result = get_line(index, number_of_elem)
    return result

